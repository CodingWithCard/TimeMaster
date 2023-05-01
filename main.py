import sys
import os
from datetime import datetime, timedelta
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

TEMP_FILE = "temp_start_time.txt"

def write_to_excel(start_time, end_time):
    file_name = "work_hours.xlsx"

    try:
        workbook = load_workbook(file_name)
    except FileNotFoundError:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append(['Date', 'Start Time', 'End Time', 'Time Difference'])

    worksheet = workbook.active

    start_datetime = datetime.strptime(start_time, "%H:%M")
    end_datetime = datetime.strptime(end_time, "%H:%M")
    total_hours = (end_datetime - start_datetime).seconds / 3600
    expected_hours = 8

    time_difference = total_hours - expected_hours
    time_difference_formatted = f"{int(time_difference):02d}:{int((time_difference * 60) % 60):02d}"
    time_difference_formatted = f"{'+' if time_difference >= 0 else '-'}{time_difference_formatted}"

    row_data = [
        datetime.now().strftime("%Y-%m-%d"),
        start_time,
        end_time,
        time_difference_formatted
    ]

    row = worksheet.append(row_data)

    last_row = worksheet.max_row
    cell_color = "00FF00" if time_difference >= 0 else "FF0000"
    worksheet.cell(row=last_row, column=4).font = Font(color=cell_color)

    workbook.save(file_name)

def start_session():
    start_time_input = input("When did you start work today? (e.g., 8:30): ")

    with open(TEMP_FILE, 'w') as temp_file:
        temp_file.write(start_time_input)

    print("Work session started.")

def stop_session():
    if not os.path.exists(TEMP_FILE):
        print("No work session found. Start a session first.")
        return

    with open(TEMP_FILE, 'r') as temp_file:
        start_time_input = temp_file.read()

    os.remove(TEMP_FILE)

    end_time_input = datetime.now().strftime("%H:%M")

    print(f"You clocked out at: {end_time_input}")

    write_to_excel(start_time_input, end_time_input)

def show_help():
    help_text = """
Usage: python work_log.py [start|stop|help]

Commands:
  start   Start a work session and record the start time.
  stop    Stop the current work session, calculate the time difference, and update the Excel file.
  help    Display this help message.
    """
    print(help_text)

def main():
    if len(sys.argv) < 2:
        show_help()
        return

    command = sys.argv[1].lower()

    if command == "start":
        start_session()
    elif command == "stop":
        stop_session()
    elif command == "help":
        show_help()
    else:
        print("Invalid command. Use 'start', 'stop', or 'help'.")

if __name__ == "__main__":
    main()
